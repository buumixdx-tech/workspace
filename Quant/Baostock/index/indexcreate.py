import baostock as bs
import pandas as pd
from datetime import datetime
import os
from openpyxl import load_workbook
from clickhouse_driver import Client

def get_index_k_data():
    print("开始执行指数数据提取...")
    
    # 登录baostock系统
    print("正在登录baostock...")
    lg = bs.login()
    print('login respond error_code:' + lg.error_code)
    print('login respond error_msg:' + lg.error_msg)

    # 检查Excel文件是否存在
    if not os.path.exists('index.xlsx'):
        print("错误：找不到index.xlsx文件")
        return
    
    print("正在读取index.xlsx文件...")
    # 使用openpyxl读取Excel文件
    try:
        wb = load_workbook('index.xlsx')
        ws = wb.active
        print("Excel文件读取成功")
    except Exception as e:
        print(f"错误：无法读取index.xlsx文件，{e}")
        return
    
    # 读取起止日期 (H2, I2)
    start_date_raw = ws['H2'].value
    end_date_raw = ws['I2'].value
    
    print(f"读取到的起始日期: {start_date_raw}")
    print(f"读取到的结束日期: {end_date_raw}")
    
    # 格式转换
    if start_date_raw is None:
        print("错误：起始日期(H2)不能为空")
        return
    
    # 转换起始日期格式 20250101 -> 2025-01-01
    try:
        start_date = datetime.strptime(str(start_date_raw), '%Y%m%d').strftime('%Y-%m-%d')
    except ValueError:
        print(f"错误：起始日期格式不正确，应为YYYYMMDD格式，当前值：{start_date_raw}")
        return
    
    # 转换截止日期格式
    if end_date_raw is None or str(end_date_raw).strip() == '':
        # 如果截止日期为空，默认为当天
        end_date = datetime.now().strftime('%Y-%m-%d')
        print(f"截止日期为空，使用默认日期：{end_date}")
    else:
        try:
            end_date = datetime.strptime(str(end_date_raw), '%Y%m%d').strftime('%Y-%m-%d')
        except ValueError:
            print(f"警告：截止日期格式不正确，使用默认日期。当前值：{end_date_raw}")
            end_date = datetime.now().strftime('%Y-%m-%d')
    
    print(f"数据提取范围：{start_date} 到 {end_date}")
    
    # 提取指数代码列表（B列，从第2行开始跳过表头）
    index_codes = []
    row_num = 2  # 从第2行开始（跳过表头）
    
    print("开始读取指数代码列表...")
    while True:
        cell_value = ws[f'B{row_num}'].value
        if cell_value is None:
            break
        if str(cell_value).strip() != '':
            index_codes.append(str(cell_value).strip())
        row_num += 1
    
    print(f"需要提取的指数数量：{len(index_codes)}")
    print(f"指数代码列表：{index_codes}")
    
    if len(index_codes) == 0:
        print("警告：未找到任何指数代码")
        return
    
    # 存储所有数据
    all_data = []
    
    # 遍历每个指数代码
    for i, code in enumerate(index_codes):
        print(f"正在提取第{i+1}个指数数据：{code}")
        
        # 获取指数日K线数据
        rs = bs.query_history_k_data_plus(
            code,
            "code,date,open,high,low,close,preclose,volume,turn,tradestatus",
            start_date=start_date,
            end_date=end_date,
            frequency="d"
        )
        
        print(f'  查询状态 error_code:{rs.error_code}, error_msg:{rs.error_msg}')
        
        if rs.error_code == '0':
            # 提取数据
            data_list = []
            while rs.next():
                row_data = rs.get_row_data()
                if len(row_data) >= 10:  # 确保数据完整
                    data_list.append(row_data)
            
            if data_list:
                all_data.extend(data_list)
                print(f'  成功提取{len(data_list)}条记录')
            else:
                print(f'  无数据')
        else:
            print(f'  查询失败：{rs.error_msg}')
    
    print(f"总共提取到 {len(all_data)} 条数据")
    
    # 转换为DataFrame并写入ClickHouse
    if all_data:
        print("正在处理数据...")
        result_df = pd.DataFrame(all_data, columns=[
            'code', 'date', 'open', 'high', 'low', 'close', 
            'preclose', 'volume', 'turn', 'tradestatus'
        ])
        
        # 数据类型转换
        print("正在进行数据类型转换...")
        numeric_columns = ['open', 'high', 'low', 'close', 'preclose', 'turn']
        for col in numeric_columns:
            result_df[col] = pd.to_numeric(result_df[col], errors='coerce').fillna(0)
        
        result_df['volume'] = pd.to_numeric(result_df['volume'], errors='coerce').fillna(0).astype('int64')
        result_df['tradestatus'] = pd.to_numeric(result_df['tradestatus'], errors='coerce').fillna(1).astype('int8')
        
        # 日期格式转换
        result_df['date'] = pd.to_datetime(result_df['date'])
        
        print(f"处理完成，准备插入 {len(result_df)} 条记录到数据库")
        print("前几行数据预览：")
        print(result_df.head())
        
        # 连接ClickHouse数据库
        try:
            print("正在连接ClickHouse数据库...")
            client = Client(
                host='localhost',
                port=9000,
                user='admin',
                password='admin_password',  
                database='default'
            )
            print("ClickHouse连接成功")
            
            # 测试连接
            test_result = client.execute("SELECT 1")
            print(f"数据库连接测试通过: {test_result}")
            
            # 插入数据到ClickHouse
            print("正在插入数据...")
            data_tuples = [tuple(row) for row in result_df.values]
            
            client.execute(
                "INSERT INTO stock_data.index_k (code, date, open, high, low, close, preclose, volume, turn, tradestatus) VALUES",
                data_tuples
            )
            
            print(f"\n数据插入完成，共插入{len(result_df)}条记录到stock_data.index_k表中")
            
            # 验证插入的数据
            count_result = client.execute("SELECT count(*) FROM stock_data.index_k")
            print(f"当前表中总记录数：{count_result[0][0]}")
            
        except Exception as e:
            print(f"错误：无法连接到ClickHouse或插入数据，{e}")
            print("请检查：")
            print("1. ClickHouse服务是否正在运行")
            print("2. host和port配置是否正确")
            print("3. 数据库stock_data是否存在")
            print("4. 表stock_data.index_k是否存在")
            # 保存到CSV作为备份
            output_file = 'index_k_data_backup.csv'
            result_df.to_csv(output_file, index=False, encoding='utf-8')
            print(f"数据已保存到备份文件：{output_file}")
    else:
        print("未提取到任何数据")
    
    # 登出系统
    bs.logout()
    print("\nbaostock连接已关闭")
    print("程序执行完成")

if __name__ == '__main__':
    get_index_k_data()
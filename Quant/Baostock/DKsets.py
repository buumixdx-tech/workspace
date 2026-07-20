import baostock as bs
import pandas as pd
import os
import time
import openpyxl
import re
from datetime import datetime

# 从data.xlsx中获取需要读取的股票列表和起止日期（H2和I2单元格），结果输出为一个xlsx，其中每个sheet都对应一只股票的日K数据。

def get_stock_data():
    try:
        print("开始获取股票数据...")
        
        # 使用openpyxl直接读取Excel文件
        wb = openpyxl.load_workbook('data.xlsx')
        sheet = wb.active
        
        # 直接从单元格读取日期信息
        start_date = sheet['H2'].value
        end_date = sheet['I2'].value
        
        # 打印日期信息
        print(f"原始起始日期: {start_date} (类型: {type(start_date)})")
        print(f"原始结束日期: {end_date} (类型: {type(end_date)})")
        
        # 处理日期格式
        def format_date(date_value):
            if date_value is None or pd.isna(date_value) or date_value == '':
                return None
                
            if isinstance(date_value, (int, float)):
                date_int = int(date_value)
                date_str = str(date_int)
                if len(date_str) == 8 and date_str.isdigit():
                    return f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}"
            
            elif isinstance(date_value, str):
                cleaned = date_value.strip()
                if len(cleaned) == 8 and cleaned.isdigit():
                    return f"{cleaned[:4]}-{cleaned[4:6]}-{cleaned[6:]}"
            
            return None
        
        start_date_str = format_date(start_date)
        end_date_str = format_date(end_date)
        
        # 打印转换后的日期
        print(f"转换后起始日期: {start_date_str}")
        print(f"转换后结束日期: {end_date_str}")
        
        # 如果结束日期为空，使用当前日期
        if not end_date_str:
            end_date_str = datetime.today().strftime('%Y-%m-%d')
            print(f"结束日期未指定，使用当前日期: {end_date_str}")
        
        # 验证起始日期
        if not start_date_str:
            print("错误：起始日期格式不正确或未指定")
            return
        
        # 获取股票代码列表
        stock_codes = []
        row_count = sheet.max_row
        
        print(f"Excel文件共有 {row_count} 行数据")
        
        # 查找股票代码所在的列索引
        stock_code_column = None
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col).value
            if cell_value and "代码" in str(cell_value):
                stock_code_column = col
                print(f"股票代码列: {openpyxl.utils.get_column_letter(col)}列")
                break
        
        if stock_code_column is None:
            print("错误：未找到股票代码列（标题可能不包含'代码'字样）")
            return
        
        # 直接读取股票代码
        print("\n读取股票代码:")
        for row in range(2, row_count + 1):  # 从第2行开始
            code_value = sheet.cell(row=row, column=stock_code_column).value
            
            # 处理可能的空值或NaN
            if code_value is None or pd.isna(code_value) or code_value == '':
                continue
                
            # 转换为字符串并处理
            code_str = str(code_value).strip()
            
            # 清理非数字字符
            cleaned_code = re.sub(r'[^0-9]', '', code_str)
            
            # 处理浮点数代码
            if '.' in cleaned_code:
                try:
                    code_float = float(cleaned_code)
                    cleaned_code = str(int(code_float))
                except:
                    pass
            
            # 填充到6位数字
            code = cleaned_code.zfill(6)
            
            if len(code) == 6 and code.isdigit():
                stock_codes.append(code)
        
        total_stocks = len(stock_codes)
        print(f"共获取 {total_stocks} 只有效股票数据")
        print(f"日期范围: {start_date_str} 至 {end_date_str}")
        
        # 登录系统
        print("正在登录baostock系统...")
        lg = bs.login()
        if lg.error_code != '0':
            print(f"登录失败: {lg.error_code} - {lg.error_msg}")
            return
        print("登录成功")
        
        # 创建Excel文件
        output_file = "dksets.xlsx"
        writer = pd.ExcelWriter(output_file, engine='xlsxwriter')
        
        # 遍历所有股票代码
        total_downloaded = 0
        for i, code in enumerate(stock_codes):
            # 确定股票市场前缀
            if code.startswith(('6', '9', '5')):
                full_code = f"sh.{code}"
            else:
                full_code = f"sz.{code}"
            
            print(f"正在下载 {full_code} ({i+1}/{total_stocks})...")
            
            # 查询历史K线数据
            rs = bs.query_history_k_data_plus(
                full_code,
                "date,code,open,high,low,close,preclose,volume,amount,adjustflag,turn,tradestatus,pctChg,isST",
                start_date=start_date_str,
                end_date=end_date_str,
                frequency="d",
                adjustflag="3"  # 后复权
            )
            
            # 检查查询结果
            if rs.error_code != '0':
                print(f"  {full_code} 查询失败: {rs.error_msg}")
                continue
            
            # 处理数据
            data_list = []
            while (rs.error_code == '0') and rs.next():
                data_list.append(rs.get_row_data())
            
            if not data_list:
                print(f"  {full_code} 未获取到数据")
                continue
            
            # 转换为DataFrame
            result = pd.DataFrame(data_list, columns=rs.fields)
            record_count = len(result)
            print(f"  获取 {record_count} 条记录")
            total_downloaded += 1
            
            # 写入Excel工作表
            result.to_excel(writer, sheet_name=code, index=False)
            
            # 避免请求过于频繁
            time.sleep(0.1)
        
        # 保存Excel文件
        writer.close()
        print(f"\n数据下载完成! 共下载 {total_downloaded}/{total_stocks} 只股票数据")
        print(f"所有数据已保存到: {os.path.abspath(output_file)}")
        
        # 登出系统
        print("登出系统")
        bs.logout()
        
    except Exception as e:
        print(f"程序出错: {str(e)}")
        # 确保登出系统
        try:
            bs.logout()
        except:
            pass
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    get_stock_data()
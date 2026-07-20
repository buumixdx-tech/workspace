import baostock as bs
import pandas as pd
from clickhouse_driver import Client
from datetime import datetime, timedelta
import concurrent.futures
import time
import logging
from threading import local
import os
from openpyxl import load_workbook
import threading

# ==================== 配置区 ====================

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('supplement_historical.log')
    ]
)
logger = logging.getLogger(__name__)

# 全局配置
HISTORICAL_START_DATE = "2018-01-01"
END_DATE = datetime.today().strftime("%Y-%m-%d")
MAX_WORKERS = 8
RETRY_LIMIT = 3

# 数据库连接配置
CH_CONFIG = {
    'host': 'localhost',
    'port': 9000,
    'user': 'admin',
    'password': 'admin_password',
    'database': 'stock_data',
    'settings': {
        'connect_timeout': 30,
        'send_receive_timeout': 120,
        'max_block_size': 10000
    }
}

EXCEL_FILE_PATH = os.path.join(os.path.dirname(__file__), "AllA.xlsx")

# ================================================
# 线程本地存储，用于管理 Baostock 会话
thread_local = local()

def get_baostock_session():
    """获取当前线程的 Baostock 登录会话（复用连接）"""
    if not hasattr(thread_local, "logged_in"):
        logger.debug(f"线程 {threading.current_thread().name} 初始化 Baostock 会话")
        ret = bs.login()
        if ret.error_code != '0':
            raise Exception(f"Baostock 登录失败: {ret.error_msg}")
        thread_local.logged_in = True
    return bs

def get_all_stock_codes():
    """从 AllA.xlsx 读取所有股票代码（使用 initial.txt 中的逻辑）"""
    if not os.path.exists(EXCEL_FILE_PATH):
        logger.error(f"Excel 文件不存在: {EXCEL_FILE_PATH}")
        return []

    try:
        wb = load_workbook(filename=EXCEL_FILE_PATH, read_only=True)
        sheet = wb.active
        stock_codes = []
        for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2):
            cell_value = str(row[0].value).strip()
            if cell_value.startswith('='):
                quoted_part = cell_value[1:].strip()
                if quoted_part.startswith('"') and quoted_part.endswith('"'):
                    code = quoted_part[1:-1].strip()
                    if code.isdigit() and len(code) == 6:
                        if code.startswith('6'):
                            stock_codes.append(f"sh.{code}")
                        elif code.startswith(('0', '3')):
                            stock_codes.append(f"sz.{code}")
        wb.close()
        logger.info(f"从 Excel 获取到 {len(stock_codes)} 只股票代码")
        return stock_codes
    except Exception as e:
        logger.error(f"读取 Excel 文件失败: {e}", exc_info=True)
        return []

def get_earliest_date_per_stock(ch_client, stock_codes):
    """
    查询每只股票在 ClickHouse 中的最早交易日期
    采用与 initial.txt 中 bulk_insert_to_clickhouse 一致的、最安全可靠的模式
    """
    logger.info("正在查询各股票在数据库中的最早交易日期...")
    earliest_dates = {}
    
    # 构建一个包含所有股票代码的 IN 查询
    # 使用 clickhouse-driver 的 execute 机制来安全地处理参数列表
    query = """
    SELECT code, MIN(date) as earliest_date
    FROM stock_data.daily_k 
    WHERE code IN ({})
    GROUP BY code
    ORDER BY code
    """.format(", ".join(["%s"] * len(stock_codes)))
    
    try:
        # ✅ 核心修复：直接将 stock_codes 列表作为参数传递给 execute
        # 这与 bulk_insert_to_clickhouse 的逻辑完全一致
        result = ch_client.execute(query, stock_codes)
        
        # 将结果转换为字典
        for row in result:
            code, earliest_date = row
            # 注意：earliest_date 可能是 None
            earliest_dates[code] = earliest_date.strftime("%Y-%m-%d") if earliest_date is not None else None
            
        # 确保所有股票都有记录（包括那些没有数据的）
        for code in stock_codes:
            if code not in earliest_dates:
                earliest_dates[code] = None
                
    except Exception as e:
        logger.error(f"批量查询所有股票最早日期失败: {e}")
        # 如果批量查询失败，为所有股票设置为 None
        for code in stock_codes:
            earliest_dates[code] = None
    
    return earliest_dates

def determine_fetch_range(stock_code, db_earliest_date):
    """
    根据数据库最早日期，决定需要补充的历史区间
    """
    if db_earliest_date is None:
        # 数据库中无此股票数据，从 HISTORICAL_START_DATE 拉到今天
        return HISTORICAL_START_DATE, END_DATE

    # 数据库中有数据，计算补充区间：从 HISTORICAL_START_DATE 到 (db_earliest_date - 1天)
    db_earliest = datetime.strptime(db_earliest_date, "%Y-%m-%d")
    historical_start = datetime.strptime(HISTORICAL_START_DATE, "%Y-%m-%d")

    if db_earliest <= historical_start:
        # 数据库最早日期已经早于或等于目标起始日，无需补充
        return None

    # 计算结束日期：db_earliest_date 的前一天
    fetch_end = (db_earliest - timedelta(days=1)).strftime("%Y-%m-%d")
    fetch_start = HISTORICAL_START_DATE

    # 确保开始日期不晚于结束日期
    if fetch_start > fetch_end:
        return None

    return fetch_start, fetch_end

def process_stock_data(df, code):
    """使用原 initial.txt 中的 robust 数据处理逻辑"""
    df = df.copy()
    df['code'] = code
    if 'date' in df.columns:
        df['date'] = pd.to_datetime(df['date']).dt.date
    type_map = {
        'open': 'float32', 'high': 'float32', 'low': 'float32',
        'close': 'float32', 'preclose': 'float32', 'turn': 'float32',
        'amount': 'float64', 'volume': 'uint64', 'tradestatus': 'uint8',
        'isST': 'uint8', 'adjustflag': 'str'
    }
    for col, dtype in type_map.items():
        if col in df.columns:
            try:
                if dtype == 'str':
                    df[col] = df[col].astype(str)
                else:
                    df[col] = pd.to_numeric(df[col], errors='coerce').astype(dtype)
            except Exception as e:
                logger.warning(f"列 {col} 转换失败: {str(e)}")
                df[col] = None
    # 处理停牌
    paused_mask = (df['tradestatus'] == 0) | (df['tradestatus'] == '0')
    df.loc[paused_mask, ['open', 'high', 'low', 'close', 'volume', 'amount']] = 0
    return df[[
        'code', 'date', 'open', 'high', 'low', 'close', 'preclose',
        'volume', 'amount', 'turn', 'tradestatus', 'isST', 'adjustflag'
    ]]

def fetch_stock_data(code, start_date, end_date, retry=0):
    """获取单只股票数据（使用线程本地会话）"""
    try:
        bs = get_baostock_session()
        rs = bs.query_history_k_data_plus(
            code=code,
            fields="date,open,high,low,close,preclose,volume,amount,turn,tradestatus,isST,adjustflag",
            start_date=start_date,
            end_date=end_date,
            frequency="d",
            adjustflag="3"  # 后复权
        )
        if rs.error_code != '0':
            raise Exception(f"API error: {rs.error_msg}")

        data_list = []
        while (row_data := rs.get_row_data()) is not None:
            data_list.append(row_data)

        df = pd.DataFrame(data_list, columns=rs.fields)
        if df.empty:
            return pd.DataFrame()

        # 使用原脚本的 robust 处理
        df = process_stock_data(df, code)
        return df

    except Exception as e:
        if retry < RETRY_LIMIT:
            time.sleep(2 * (retry + 1))
            return fetch_stock_data(code, start_date, end_date, retry + 1)
        logger.error(f"获取 {code} [{start_date} 到 {end_date}] 数据失败: {e}")
        return pd.DataFrame()

def bulk_insert_to_clickhouse(ch_client, df):
    """批量写入 ClickHouse"""
    if df.empty:
        return
    try:
        df['date'] = pd.to_datetime(df['date']).dt.date
        data = [tuple(None if pd.isna(x) else x for x in row) for row in df.values]
        batch_size = 5000
        for i in range(0, len(data), batch_size):
            batch = data[i:i+batch_size]
            # 保持与 initial.txt 一致的写入方式 - 不使用参数化查询
            ch_client.execute(
                '''INSERT INTO stock_data.daily_k (
                    code, date, open, high, low, close, preclose,
                    volume, amount, turn, tradestatus, isST, adjustflag
                ) VALUES''',
                batch,
                settings={'input_format_null_as_default': 1}
            )
        logger.info(f"成功写入 {len(df)} 行数据")
    except Exception as e:
        logger.error(f"写入 ClickHouse 失败: {e}")
        raise

def process_single_stock(args):
    """
    处理单只股票的补充任务
    args: (code, fetch_start, fetch_end)
    """
    code, fetch_start, fetch_end = args
    logger.info(f"开始补充 {code} 的历史数据: {fetch_start} 到 {fetch_end}")

    # 在任务内部创建自己的 ClickHouse 连接
    ch_client = Client(**CH_CONFIG)
    try:
        df = fetch_stock_data(code, fetch_start, fetch_end)
        if df.empty:
            logger.info(f"{code} 在 {fetch_start} 到 {fetch_end} 无数据")
            return code, True, 0

        bulk_insert_to_clickhouse(ch_client, df)
        return code, True, len(df)

    except Exception as e:
        logger.error(f"{code} 补充历史数据失败: {e}", exc_info=True)
        return code, False, 0
    finally:
        ch_client.disconnect()  # 确保连接关闭

def main_supplement():
    """主函数：补充历史数据"""
    start_time = time.time()
    logger.info(f"开始补充历史数据，目标起始日期: {HISTORICAL_START_DATE}")

    # 1. 初始化 ClickHouse 客户端
    ch_client = Client(**CH_CONFIG)
    try:
        # 2. 获取所有股票代码
        all_codes = get_all_stock_codes()
        if not all_codes:
            logger.error("未能获取股票代码，终止任务")
            return

        # 3. 查询每只股票在数据库中的最早日期
        earliest_dates = get_earliest_date_per_stock(ch_client, all_codes)
        ch_client.disconnect()  # 关闭初始查询连接

        # 4. 确定需要拉取的任务列表
        tasks = []
        for code in all_codes:
            db_earliest = earliest_dates[code]
            fetch_range = determine_fetch_range(code, db_earliest)
            if fetch_range:
                start, end = fetch_range
                tasks.append((code, start, end))

        logger.info(f"共需补充 {len(tasks)} 只股票的历史数据")

        if not tasks:
            logger.info("所有股票的历史数据均已满足要求，无需补充。")
            return

        # 5. 并发执行补充任务
        success_count = 0
        fail_count = 0
        total_rows_inserted = 0

        with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS, thread_name_prefix="SuppWorker") as executor:
            futures = [executor.submit(process_single_stock, task) for task in tasks]

            for future in concurrent.futures.as_completed(futures):
                code, success, rows = future.result()
                if success:
                    success_count += 1
                    total_rows_inserted += rows
                else:
                    fail_count += 1

        # 6. 输出结果报告
        duration = time.time() - start_time
        logger.info(f"""
=== 历史数据补充完成 ===
耗时: {duration:.2f} 秒
成功率: {success_count}/{len(tasks)} ({success_count/len(tasks)*100:.1f}%)
总写入行数: {total_rows_inserted}
失败任务: {fail_count}
        """)

    except Exception as e:
        logger.critical(f"主流程发生严重错误: {e}", exc_info=True)

if __name__ == "__main__":
    main_supplement()
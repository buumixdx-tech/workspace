import baostock as bs
import pandas as pd
from clickhouse_driver import Client
from datetime import datetime, timedelta
import time
import logging
from openpyxl import load_workbook
import os

# ==================== 配置区 ====================

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('supplement_historical.log')
    ]
)
logger = logging.getLogger(__name__)

# 全局配置
HISTORICAL_START_DATE = "2025-05-01"
END_DATE = datetime.today().strftime("%Y-%m-%d")
RETRY_LIMIT = 2  # 与 initial.txt 一致

# 数据库连接配置
CH_CONFIG = {
    'host': 'localhost',
    'port': 9000,
    'user': 'admin',
    'password': 'admin_password',
    'database': 'stock_data',
    'settings': {
        'connect_timeout': 30,
        'send_receive_timeout': 60
    }
}

EXCEL_FILE_PATH = os.path.join(os.path.dirname(__file__), "AllA.xlsx")

# ================================================

def get_all_stock_codes():
    """从 AllA.xlsx 读取所有股票代码（使用 initial.txt 中的逻辑）"""
    if not os.path.exists(EXCEL_FILE_PATH):
        logger.error(f"Excel 文件不存在: {EXCEL_FILE_PATH}")
        return []

    try:
        wb = load_workbook(filename=EXCEL_FILE_PATH, read_only=True)
        sheet = wb.active
        stock_codes = []
        for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2):
            cell_value = str(row[0].value).strip()
            if cell_value.startswith('='):
                quoted_part = cell_value[1:].strip()
                if quoted_part.startswith('"') and quoted_part.endswith('"'):
                    code = quoted_part[1:-1].strip()
                    if code.isdigit() and len(code) == 6:
                        if code.startswith('6'):
                            stock_codes.append(f"sh.{code}")
                        elif code.startswith(('0', '3')):
                            stock_codes.append(f"sz.{code}")
        wb.close()
        logger.info(f"从 Excel 获取到 {len(stock_codes)} 只股票代码")
        return stock_codes
    except Exception as e:
        logger.error(f"读取 Excel 文件失败: {e}", exc_info=True)
        return []

def get_earliest_date_per_stock(ch_client, stock_codes):
    """
    查询每只股票在 ClickHouse 中的最早交易日期
    采用与 initial.txt 中 main_initial_import 一致的串行模式
    """
    logger.info("正在查询各股票在数据库中的最早交易日期...")
    earliest_dates = {}
    
    # 串行查询，与 initial.txt 保持一致
    for code in stock_codes:
        try:
            # 直接构建 SQL 查询，不使用参数化
            query = f"SELECT MIN(date) FROM stock_data.daily_k WHERE code = '{code}'"
            result = ch_client.execute(query)
            earliest_date = result[0][0] if result and result[0][0] else None
            earliest_dates[code] = earliest_date.strftime("%Y-%m-%d") if earliest_date else None
        except Exception as e:
            logger.warning(f"查询 {code} 的最早日期失败，视为无数据: {e}")
            earliest_dates[code] = None
    
    return earliest_dates

def determine_fetch_range(stock_code, db_earliest_date):
    """
    根据数据库最早日期，决定需要补充的历史区间
    """
    if db_earliest_date is None:
        return HISTORICAL_START_DATE, END_DATE

    db_earliest = datetime.strptime(db_earliest_date, "%Y-%m-%d")
    historical_start = datetime.strptime(HISTORICAL_START_DATE, "%Y-%m-%d")

    if db_earliest <= historical_start:
        return None

    fetch_end = (db_earliest - timedelta(days=1)).strftime("%Y-%m-%d")
    fetch_start = HISTORICAL_START_DATE

    if fetch_start > fetch_end:
        return None

    return fetch_start, fetch_end

def process_stock_data(df, code):
    """使用原 initial.txt 中的 robust 数据处理逻辑"""
    df = df.copy()
    df['code'] = code
    if 'date' in df.columns:
        df['date'] = pd.to_datetime(df['date']).dt.date
    type_map = {
        'open': 'float32', 'high': 'float32', 'low': 'float32',
        'close': 'float32', 'preclose': 'float32', 'turn': 'float32',
        'amount': 'float64', 'volume': 'uint64', 'tradestatus': 'uint8',
        'isST': 'uint8', 'adjustflag': 'str'
    }
    for col, dtype in type_map.items():
        if col in df.columns:
            try:
                if dtype == 'str':
                    df[col] = df[col].astype(str)
                else:
                    df[col] = pd.to_numeric(df[col], errors='coerce').astype(dtype)
            except Exception as e:
                logger.warning(f"列 {col} 转换失败: {str(e)}")
                df[col] = None
    paused_mask = (df['tradestatus'] == 0) | (df['tradestatus'] == '0')
    df.loc[paused_mask, ['open', 'high', 'low', 'close', 'volume', 'amount']] = 0
    return df[[
        'code', 'date', 'open', 'high', 'low', 'close', 'preclose',
        'volume', 'amount', 'turn', 'tradestatus', 'isST', 'adjustflag'
    ]]

def fetch_stock_data(code, start_date, end_date, retry=0):
    """获取单只股票数据（与 initial.txt 完全一致）"""
    try:
        bs.login()
        rs = bs.query_history_k_data_plus(
            code=code,
            fields="date,open,high,low,close,preclose,volume,amount,turn,tradestatus,isST,adjustflag",
            start_date=start_date,
            end_date=end_date,
            frequency="d",
            adjustflag="3"
        )
        if rs.error_code != '0':
            logger.warning(f"股票 {code} 查询失败: {rs.error_msg}")
            return pd.DataFrame()

        data_list = []
        while rs.next():
            data_list.append(rs.get_row_data())

        df = pd.DataFrame(data_list, columns=rs.fields)
        bs.logout()
        
        if not df.empty:
            df = process_stock_data(df, code)
            return df
        
        return pd.DataFrame()

    except Exception as e:
        if retry < RETRY_LIMIT:
            time.sleep(2)
            return fetch_stock_data(code, start_date, end_date, retry + 1)
        logger.error(f"股票 {code} 获取失败(重试{RETRY_LIMIT}次): {str(e)}")
        return pd.DataFrame()

def bulk_insert_to_clickhouse(ch_client, df):
    """批量写入 ClickHouse（与 initial.txt 保持一致）"""
    if df.empty:
        return
    try:
        data = []
        for _, row in df.iterrows():
            data.append(tuple(None if pd.isna(x) else x for x in row.values))
        
        batch_size = 10000
        for i in range(0, len(data), batch_size):
            batch = data[i:i+batch_size]
            ch_client.execute(
                '''INSERT INTO stock_data.daily_k (
                    code, date, open, high, low, close, preclose,
                    volume, amount, turn, tradestatus, isST, adjustflag
                ) VALUES''',
                batch,
                settings={'input_format_null_as_default': 1}
            )
            logger.info(f"已写入批次 {i//batch_size + 1} (行 {i}-{i+len(batch)})")
    except Exception as e:
        logger.error(f"写入失败: {str(e)}")
        raise

def main_supplement():
    """主函数：补充历史数据（串行模式，与 initial.txt 一致）"""
    start_time = time.time()
    logger.info(f"开始补充历史数据，目标起始日期: {HISTORICAL_START_DATE}")

    # 1. 初始化 ClickHouse 客户端
    ch_client = Client(**CH_CONFIG)
    
    # 2. 获取所有股票代码
    all_codes = get_all_stock_codes()
    if not all_codes:
        logger.error("未能获取股票代码，终止任务")
        return

    # 3. 串行查询每只股票的最早日期
    # 严格遵循 initial.txt 的串行模式
    logger.info("正在串行查询各股票的最早交易日期...")
    earliest_dates = {}
    for idx, code in enumerate(all_codes, 1):
        try:
            query = f"SELECT MIN(date) FROM stock_data.daily_k WHERE code = '{code}'"
            result = ch_client.execute(query)
            earliest_date = result[0][0] if result and result[0][0] else None
            earliest_dates[code] = earliest_date.strftime("%Y-%m-%d") if earliest_date else None
            logger.debug(f"[{idx}/{len(all_codes)}] 查询 {code} 完成")
        except Exception as e:
            logger.warning(f"[{idx}/{len(all_codes)}] 查询 {code} 失败: {e}")
            earliest_dates[code] = None

    # 4. 确定需要拉取的任务列表
    tasks = []
    for code in all_codes:
        db_earliest = earliest_dates[code]
        fetch_range = determine_fetch_range(code, db_earliest)
        if fetch_range:
            start, end = fetch_range
            tasks.append((code, start, end))

    logger.info(f"共需补充 {len(tasks)} 只股票的历史数据")

    if not tasks:
        logger.info("所有股票的历史数据均已满足要求，无需补充。")
        return

    # 5. 串行执行补充任务
    success_count = 0
    fail_count = 0

    for idx, (code, fetch_start, fetch_end) in enumerate(tasks, 1):
        logger.info(f"[{idx}/{len(tasks)}] 开始补充 {code} 的历史数据: {fetch_start} 到 {fetch_end}")
        try:
            df = fetch_stock_data(code, fetch_start, fetch_end)
            if df.empty:
                logger.info(f"{code} 在 {fetch_start} 到 {fetch_end} 无数据")
                fail_count += 1
                continue

            bulk_insert_to_clickhouse(ch_client, df)
            success_count += 1
        except Exception as e:
            logger.error(f"[{idx}/{len(tasks)}] {code} 补充历史数据失败: {e}", exc_info=True)
            fail_count += 1

    # 6. 输出结果报告
    duration = time.time() - start_time
    logger.info(f"""
=== 历史数据补充完成 ===
耗时: {duration:.2f} 秒
成功率: {success_count}/{len(tasks)} ({success_count/len(tasks)*100:.1f}%)
失败任务: {fail_count}
        """)

    ch_client.disconnect()

if __name__ == "__main__":
    main_supplement()
import baostock as bs
import pandas as pd
from concurrent.futures import ThreadPoolExecutor, as_completed
import time
import logging
from openpyxl import load_workbook
import os
from datetime import datetime
from queue import Queue
import threading

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('test_update.log')
    ]
)
logger = logging.getLogger(__name__)

# 全局配置
CONCURRENT_WORKERS = 16  # 并发线程数
RETRY_LIMIT = 1  # 降低重试次数，因为登录不再频繁
TEST_START_DATE = '2025-09-11'
TEST_END_DATE = '2025-09-12'

# ======== 新增：全局锁和登录状态 ========
login_lock = threading.Lock()  # 保护登录操作的锁
is_logged_in = False  # 全局登录状态标志
lg = None  # 全局登录对象

def ensure_login():
    """确保当前线程拥有一个有效的登录会话"""
    global is_logged_in, lg
    with login_lock:
        if not is_logged_in:
            logger.info("正在初始化全局登录...")
            lg = bs.login()
            if lg.error_code != '0':
                logger.error(f"全局登录BaoStock失败: {lg.error_msg}")
                raise Exception(f"全局登录失败: {lg.error_msg}")
            else:
                is_logged_in = True
                logger.info("全局登录成功！")

def fetch_single_stock_data(stock_code, start_date, end_date, retry=0):
    """获取单只股票数据 - 使用全局登录"""
    try:
        # 确保登录
        ensure_login()
        
        rs = bs.query_history_k_data_plus(
            code=stock_code,
            fields="date,code,open,high,low,close,preclose,volume,amount,turn,tradestatus,isST,adjustflag",
            start_date=start_date,
            end_date=end_date,
            frequency="d",
            adjustflag="3"  # 后复权
        )
        data_list = []
        if rs.error_code != '0':
            logger.warning(f"股票 {stock_code} 查询失败: {rs.error_msg}")
        else:
            while rs.next():
                data_list.append(rs.get_row_data())
        # 注意：这里不调用 bs.logout()！保持会话活跃
        
        if not data_list:
            return pd.DataFrame()
        columns = rs.fields
        df = pd.DataFrame(data_list, columns=columns)
        # 简单数据处理
        numeric_cols = ['open', 'high', 'low', 'close', 'preclose', 'turn', 'amount']
        df[numeric_cols] = df[numeric_cols].apply(pd.to_numeric, errors='coerce')
        df['volume'] = pd.to_numeric(df['volume'], errors='coerce').fillna(0).astype('uint64')
        df['tradestatus'] = df['tradestatus'].astype('uint8')
        df['isST'] = df['isST'].astype('uint8')
        # 处理停牌日数据
        df.loc[df['tradestatus'] == 0, ['open', 'high', 'low', 'close', 'volume', 'amount']] = 0
        return df
    except Exception as e:
        if retry < RETRY_LIMIT:
            time.sleep(2) # 等待后重试
            return fetch_single_stock_data(stock_code, start_date, end_date, retry+1)
        logger.error(f"获取股票 {stock_code} 数据失败(重试{RETRY_LIMIT}次): {str(e)}")
        return pd.DataFrame()

def fetch_stock_data_concurrent(stock_codes, start_date, end_date, max_workers):
    """并发获取多只股票数据"""
    all_data = []
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_code = {
            executor.submit(fetch_single_stock_data, code, start_date, end_date): code 
            for code in stock_codes
        }
        for future in as_completed(future_to_code):
            code = future_to_code[future]
            try:
                df = future.result()
                if not df.empty:
                    all_data.append(df)
                    logger.info(f"成功获取 {code} 的数据，共 {len(df)} 条")
            except Exception as e:
                logger.error(f"处理 {code} 时发生异常: {str(e)}")
    # 合并所有数据
    if all_data:
        return pd.concat(all_data, ignore_index=True)
    else:
        return pd.DataFrame()

def test_concurrent_performance():
    """测试并发性能"""
    logger.info("="*60)
    logger.info("开始测试Baostock并发查询性能")
    logger.info(f"测试日期范围: {TEST_START_DATE} 到 {TEST_END_DATE}")
    logger.info(f"并发线程数: {CONCURRENT_WORKERS}")
    logger.info("="*60)
    
    # 获取所有股票代码
    all_codes = get_all_stock_codes()
    if not all_codes:
        logger.error("无法获取股票代码列表，终止测试")
        return
    
    logger.info(f"共获取 {len(all_codes)} 只股票代码，开始测试...")
    
    start_time = time.time()
    result_df = fetch_stock_data_concurrent(all_codes, TEST_START_DATE, TEST_END_DATE, CONCURRENT_WORKERS)
    end_time = time.time()
    duration = end_time - start_time
    
    total_records = len(result_df)
    unique_stocks = result_df['code'].nunique() if not result_df.empty else 0
    
    logger.info("="*60)
    logger.info("测试结果:")
    logger.info(f"1. 共采集了 {total_records} 条数据")
    logger.info(f"2. 采集对象是 {unique_stocks} 个个股")
    logger.info(f"3. 总计用时 {duration:.2f} 秒")
    logger.info(f"4. 平均每秒处理 {total_records/duration:.2f} 条数据")
    logger.info(f"5. 平均每个股票用时 {duration/len(all_codes):.4f} 秒")
    logger.info("="*60)
    
    if not result_df.empty:
        result_df.to_csv(f"test_result_{CONCURRENT_WORKERS}workers.csv", index=False)
        logger.info(f"测试结果已保存到 test_result_{CONCURRENT_WORKERS}workers.csv")

def get_all_stock_codes():
    """从AllA.xlsx获取所有沪深A股代码（处理="XXXXXX"格式）"""
    try:
        file_path = os.path.join(os.path.dirname(__file__), "AllA.xlsx")
        if not os.path.exists(file_path):
            logger.error(f"文件不存在: {file_path}")
            return []
        
        logger.info(f"正在从 {file_path} 读取股票代码...")
        wb = load_workbook(filename=file_path, read_only=True)
        sheet = wb.active
        stock_codes = []
        
        for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2):
            cell_value = str(row[0].value).strip()
            if cell_value.startswith('='):
                quoted_part = cell_value[1:].strip()
                if quoted_part.startswith('"') and quoted_part.endswith('"'):
                    code = quoted_part[1:-1].strip()
                    if code.isdigit() and len(code) == 6:
                        if code.startswith('6'):
                            stock_codes.append(f"sh.{code}")
                        elif code.startswith(('0', '3')):
                            stock_codes.append(f"sz.{code}")
        
        wb.close()
        logger.info(f"成功读取 {len(stock_codes)} 只股票代码")
        return stock_codes
    except Exception as e:
        logger.error(f"读取股票代码失败: {str(e)}", exc_info=True)
        return []

if __name__ == "__main__":
    test_concurrent_performance()
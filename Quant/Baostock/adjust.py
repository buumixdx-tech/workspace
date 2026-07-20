import baostock as bs
import pandas as pd
import logging
from clickhouse_driver import Client
from datetime import datetime
import time
import os
import concurrent.futures
from openpyxl import load_workbook
import threading
from queue import Queue

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class AdjustFactorDownloader:
    def __init__(self, excel_file_path='AllA.xlsx', max_workers=10):
        # ClickHouse连接配置
        self.ch_client = Client(
            host='localhost',
            port=9000,
            user='admin',
            password='admin_password',
            database='stock_data'
        )
        self.excel_file_path = excel_file_path
        self.max_workers = max_workers
        self.lock = threading.Lock()
        self.processed_count = 0
        self.total_stocks = 0
        
    def get_all_stock_codes(self):
        """从AllA.xlsx获取所有沪深A股代码（同update.py）"""
        try:
            file_path = os.path.join(os.path.dirname(__file__), "AllA.xlsx")
            if not os.path.exists(file_path):
                logger.error(f"文件不存在: {file_path}")
                return []
            
            logger.info(f"正在从 {file_path} 读取股票代码...")
            wb = load_workbook(filename=file_path, read_only=True)
            sheet = wb.active
            stock_codes = []
            
            for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2):
                cell_value = str(row[0].value).strip()
                if cell_value.startswith('='):
                    quoted_part = cell_value[1:].strip()
                    if quoted_part.startswith('"') and quoted_part.endswith('"'):
                        code = quoted_part[1:-1].strip()
                        if code.isdigit() and len(code) == 6:
                            if code.startswith('6'):
                                stock_codes.append(f"sh.{code}")
                            elif code.startswith(('0', '3')):
                                stock_codes.append(f"sz.{code}")
            
            wb.close()
            logger.info(f"成功读取 {len(stock_codes)} 只股票代码")
            return stock_codes
        except Exception as e:
            logger.error(f"读取股票代码失败: {str(e)}", exc_info=True)
            return []
    
    def download_adjust_factors(self, full_code, start_date, end_date):
        """下载单只股票的复权因子"""
        try:
            # 查询复权因子
            rs_factor = bs.query_adjust_factor(
                code=full_code, 
                start_date=start_date, 
                end_date=end_date
            )
            
            data_list = []
            while (rs_factor.error_code == '0') and rs_factor.next():
                data_list.append(rs_factor.get_row_data())
            
            if data_list:
                result = pd.DataFrame(data_list, columns=rs_factor.fields)
                # 添加数据更新时间
                result['dt'] = datetime.now()
                return result
            else:
                return pd.DataFrame()
                
        except Exception as e:
            logger.error(f"下载 {full_code} 复权因子失败: {e}")
            return pd.DataFrame()
    
    def insert_to_clickhouse(self, df):
        """插入数据到ClickHouse"""
        if df.empty:
            return
        
        try:
            # 转换数据类型
            df['dividOperateDate'] = pd.to_datetime(df['dividOperateDate']).dt.date
            df['foreAdjustFactor'] = pd.to_numeric(df['foreAdjustFactor'], errors='coerce')
            df['backAdjustFactor'] = pd.to_numeric(df['backAdjustFactor'], errors='coerce')
            
            # 移除空值
            df = df.dropna(subset=['foreAdjustFactor', 'backAdjustFactor'])
            
            if not df.empty:
                # 转换为元组列表用于插入
                data = [
                    (
                        row['code'],  # 这里保持完整的sh./sz.前缀
                        row['dividOperateDate'],
                        float(row['foreAdjustFactor']),
                        float(row['backAdjustFactor']),
                        row['dt']
                    )
                    for _, row in df.iterrows()
                ]
                
                # 插入数据
                insert_sql = """
                INSERT INTO adjust_factor (code, dividOperateDate, foreAdjustFactor, backAdjustFactor, dt)
                VALUES
                """
                self.ch_client.execute(insert_sql, data)
                return len(data)
            else:
                return 0
                
        except Exception as e:
            logger.error(f"插入数据到ClickHouse失败: {e}")
            return 0
    
    def process_single_stock(self, full_code, start_date, end_date):
        """处理单只股票的下载和插入"""
        try:
            # 下载复权因子
            df_factors = self.download_adjust_factors(full_code, start_date, end_date)
            
            if not df_factors.empty:
                # 插入到ClickHouse
                inserted_count = self.insert_to_clickhouse(df_factors)
                
                # 更新计数
                with self.lock:
                    self.processed_count += 1
                    progress = (self.processed_count / self.total_stocks) * 100
                
                logger.info(f"{full_code} 完成 ({self.processed_count}/{self.total_stocks}, {progress:.1f}%) - 获取 {len(df_factors)} 条记录")
                return True, len(df_factors)
            else:
                with self.lock:
                    self.processed_count += 1
                    progress = (self.processed_count / self.total_stocks) * 100
                
                logger.info(f"{full_code} 完成 ({self.processed_count}/{self.total_stocks}, {progress:.1f}%) - 无数据")
                return True, 0
                
        except Exception as e:
            logger.error(f"处理 {full_code} 时出错: {e}")
            with self.lock:
                self.processed_count += 1
            return False, 0
    
    def run_with_threadpool(self, stock_codes, start_date, end_date):
        """使用线程池并行处理"""
        self.total_stocks = len(stock_codes)
        successful_count = 0
        total_records = 0
        
        # 使用线程池执行器
        with concurrent.futures.ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            # 提交所有任务
            future_to_code = {
                executor.submit(self.process_single_stock, code, start_date, end_date): code 
                for code in stock_codes
            }
            
            # 处理完成的任务
            for future in concurrent.futures.as_completed(future_to_code):
                code = future_to_code[future]
                try:
                    success, records = future.result()
                    if success:
                        successful_count += 1
                        total_records += records
                except Exception as e:
                    logger.error(f"任务 {code} 执行异常: {e}")
        
        return successful_count, total_records
    
    def run_with_batch_threading(self, stock_codes, start_date, end_date):
        """使用批量线程处理（更细粒度的控制）"""
        self.total_stocks = len(stock_codes)
        successful_count = 0
        total_records = 0
        
        # 分批处理
        batch_size = self.max_workers * 2
        total_batches = (len(stock_codes) + batch_size - 1) // batch_size
        
        for batch_idx in range(total_batches):
            start_idx = batch_idx * batch_size
            end_idx = min((batch_idx + 1) * batch_size, len(stock_codes))
            batch_codes = stock_codes[start_idx:end_idx]
            
            logger.info(f"处理批次 {batch_idx + 1}/{total_batches}, 本批 {len(batch_codes)} 只股票")
            
            threads = []
            results = Queue()
            
            def worker(code):
                try:
                    df_factors = self.download_adjust_factors(code, start_date, end_date)
                    if not df_factors.empty:
                        records_count = self.insert_to_clickhouse(df_factors)
                        results.put((True, records_count))
                    else:
                        results.put((True, 0))
                except Exception as e:
                    logger.error(f"线程处理 {code} 失败: {e}")
                    results.put((False, 0))
            
            # 创建并启动线程
            for code in batch_codes:
                thread = threading.Thread(target=worker, args=(code,))
                thread.start()
                threads.append(thread)
                time.sleep(0.05)  # 稍微延迟避免同时发起太多请求
            
            # 等待所有线程完成
            for thread in threads:
                thread.join()
            
            # 统计本批次结果
            batch_success = 0
            batch_records = 0
            while not results.empty():
                success, records = results.get()
                if success:
                    batch_success += 1
                    batch_records += records
            
            successful_count += batch_success
            total_records += batch_records
            
            logger.info(f"批次 {batch_idx + 1} 完成: {batch_success}/{len(batch_codes)} 成功, {batch_records} 条记录")
        
        return successful_count, total_records
    
    def run(self):
        """主运行函数"""
        start_date = "2020-01-01"
        end_date = "2025-08-22"
        
        logger.info("开始登录baostock系统...")
        lg = bs.login()
        if lg.error_code != '0':
            logger.error(f"登录失败: {lg.error_code} - {lg.error_msg}")
            return
        
        logger.info("登录成功")
        
        try:
            # 从Excel文件获取所有股票代码
            stock_codes = self.get_all_stock_codes()
            if not stock_codes:
                logger.error("未获取到股票代码，程序退出")
                return
            
            total_stocks = len(stock_codes)
            logger.info(f"开始下载 {total_stocks} 只股票的复权因子数据，使用 {self.max_workers} 个线程...")
            
            start_time = time.time()
            
            # 使用方法1：线程池执行器（推荐）
            successful_count, total_records = self.run_with_threadpool(stock_codes, start_date, end_date)
            
            # 或者使用方法2：批量线程处理
            # successful_count, total_records = self.run_with_batch_threading(stock_codes, start_date, end_date)
            
            end_time = time.time()
            elapsed_time = end_time - start_time
            
            logger.info(f"下载完成！")
            logger.info(f"成功处理: {successful_count}/{total_stocks} 只股票")
            logger.info(f"总记录数: {total_records} 条")
            logger.info(f"总耗时: {elapsed_time:.2f} 秒")
            logger.info(f"平均每只股票: {elapsed_time/total_stocks:.2f} 秒")
            
        except Exception as e:
            logger.error(f"程序执行出错: {e}")
            import traceback
            traceback.print_exc()
        
        finally:
            # 登出系统
            bs.logout()
            logger.info("已登出baostock系统")
            
            # 关闭ClickHouse连接
            self.ch_client.disconnect()
            logger.info("已关闭ClickHouse连接")

if __name__ == "__main__":
    # 检查必要的库
    try:
        import baostock
        import clickhouse_driver
    except ImportError as e:
        print(f"请安装必要的库: pip install baostock clickhouse-driver pandas openpyxl")
        exit(1)
    
    # 配置参数
    excel_file_path = 'AllA.xlsx'  # Excel文件路径
    max_workers = 15  # 线程数，根据网络情况和服务器承受能力调整
    
    # 运行下载器
    downloader = AdjustFactorDownloader(
        excel_file_path=excel_file_path,
        max_workers=max_workers
    )
    downloader.run()
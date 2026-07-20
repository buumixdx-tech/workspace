# sampling.py
import os
import subprocess
import sys
import pandas as pd
from datetime import datetime
import flu2  # 导入 flu2.py 以复用其函数
from openpyxl import load_workbook # 需要导入 load_workbook 来调整列宽

# --- 工具函数 ---
def get_sampling_config(config_path: str = 'sampling_config.txt') -> dict:
    """读取回测配置文件"""
    config = {}
    if os.path.exists(config_path):
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    if not line or ':' not in line:
                        continue
                    parts = line.split(':', 1)
                    key = parts[0].strip()
                    value_raw = parts[1].strip()
                    if value_raw:
                        # 简单处理布尔值
                        if key == 'filter':
                            config[key] = value_raw.lower() in ['true', '1', 'yes']
                        elif key == 'BackTestingMode':
                             try:
                                 config[key] = int(value_raw)
                             except ValueError:
                                 print(f"警告：BackTestingMode '{value_raw}' 不是有效整数，将被忽略。")
                                 config[key] = None
                        else:
                            config[key] = value_raw
        except Exception as e:
            print(f"读取回测配置文件 {config_path} 时发生错误: {e}")
    return config

def update_config_file(config_path: str, end_date_str: str):
    """修改 flu2.py 的配置文件 config.txt 中的 end_date"""
    config_lines = []
    end_date_updated = False
    print(f"  更新 {config_path} 中的 end_date 为 {end_date_str}...")
    if os.path.exists(config_path):
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                for line in f:
                    if line.strip().startswith('end_date:'):
                        config_lines.append(f"end_date:{end_date_str}\n") # Ensure newline
                        end_date_updated = True
                    else:
                        config_lines.append(line)
        except Exception as e:
            print(f"读取配置文件 {config_path} 时出错: {e}")
            return False
    else:
        print(f"配置文件 {config_path} 不存在，将创建。")
        config_lines = []
    if not end_date_updated:
        config_lines.append(f"end_date:{end_date_str}\n") # Ensure newline
    try:
        with open(config_path, 'w', encoding='utf-8') as f:
            f.writelines(config_lines)
        print(f"  {config_path} 已更新。")
        return True
    except Exception as e:
        print(f"写入配置文件 {config_path} 时出错: {e}")
        return False

def run_strategy_script(script_name: str = 'flu2.py'):
    """执行策略脚本"""
    print(f"  执行策略脚本 {script_name}...")
    try:
        # 使用 subprocess 运行 flu2.py
        # sys.executable 确保使用当前 Python 解释器
        result = subprocess.run([sys.executable, script_name], check=True, capture_output=True, text=True)
        print("  策略脚本执行成功。")
        # Optional: Print output for debugging
        # print("STDOUT:", result.stdout)
        # if result.stderr:
        #     print("STDERR:", result.stderr)
        return True
    except subprocess.CalledProcessError as e:
        print(f"  执行策略脚本 {script_name} 时出错 (退出码 {e.returncode}): {e}")
        # Optional: Print output for debugging
        # if e.stdout:
        #     print(f"    STDOUT: {e.stdout}")
        # if e.stderr:
        #     print(f"    STDERR: {e.stderr}")
        return False
    except FileNotFoundError:
        print(f"  未找到策略脚本文件: {script_name}")
        return False
    except Exception as e:
        print(f"  执行策略脚本 {script_name} 时发生未知错误: {e}")
        return False

def read_strategy_results(result_file: str = 'result.xlsx', sheet_name: str = 'results') -> pd.DataFrame:
    """读取策略执行结果"""
    if not os.path.exists(result_file):
        print(f"  未找到策略结果文件: {result_file}")
        return pd.DataFrame()
    try:
        df_results = pd.read_excel(result_file, sheet_name=sheet_name)
        print(f"  成功读取 {result_file} 的 '{sheet_name}' sheet。")
        # 假设 A 列是 '股票代码' (根据 flu2.py 输出)
        if '股票代码' in df_results.columns:
             # 只保留 '股票代码' 列
             return df_results[['股票代码']].copy()
        else:
            print(f"  '{sheet_name}' sheet 中未找到 '股票代码' 列。")
            return pd.DataFrame()
    except Exception as e:
        print(f"  读取策略结果文件 {result_file} 时出错: {e}")
        return pd.DataFrame()

def filter_results(df_all_results: pd.DataFrame, trading_calendar: list) -> pd.DataFrame:
    """对所有选股结果进行二次筛选"""
    if df_all_results.empty:
        print("无数据可供筛选。")
        return df_all_results

    print("开始执行二次筛选...")
    filtered_data = []

    # 按交易日排序处理
    df_all_results_sorted = df_all_results.sort_values(by='交易日')
    
    # 创建一个字典，方便快速查找某日的股票集合
    date_to_stocks_set = {date: set(group['股票代码']) 
                          for date, group in df_all_results_sorted.groupby('交易日')}
    
    # 创建一个字典，方便快速查找某日的原始 DataFrame (包含可能的 NaN)
    date_to_original_df = {date: group 
                           for date, group in df_all_results_sorted.groupby('交易日')}

    # 按交易日排序处理
    sorted_dates = sorted(date_to_stocks_set.keys())
    
    for i, current_date in enumerate(sorted_dates):
        current_stocks_set = date_to_stocks_set[current_date]
        # 如果是最后一个交易日，不进行对比，直接保留
        if i == len(sorted_dates) - 1:
            print(f"  日期 {current_date}: 最后一个交易日，不进行筛选，保留所有股票。")
            # 保留原始 DataFrame
            df_current = date_to_original_df[current_date].copy()
            filtered_data.append(df_current)
            continue

        # 获取下一个交易日
        next_date = sorted_dates[i + 1]
        next_stocks_set = date_to_stocks_set[next_date]

        # 找出重叠的股票代码
        overlap_codes = current_stocks_set.intersection(next_stocks_set)

        if overlap_codes:
            print(f"  日期 {current_date}: 发现重叠股票 {list(overlap_codes)}，将被移除。")
            # 过滤掉重叠的股票，保留非重叠的
            # 从原始 DataFrame 中过滤，以保留所有行（包括可能的 NaN）
            df_filtered_day = date_to_original_df[current_date][
                ~date_to_original_df[current_date]['股票代码'].isin(overlap_codes)
            ].copy()
        else:
            print(f"  日期 {current_date}: 无重叠股票。")
            # 保留原始 DataFrame
            df_filtered_day = date_to_original_df[current_date].copy()
            
        filtered_data.append(df_filtered_day)

    if filtered_data:
        df_filtered_final = pd.concat(filtered_data, ignore_index=True)
        print("二次筛选完成。")
        return df_filtered_final
    else:
        print("筛选后无数据。")
        # 如果筛选后为空，返回一个空的 DataFrame，但保持列结构
        return pd.DataFrame(columns=['交易日', '股票代码'])

# --- 辅助函数：将长格式 DataFrame 转换为宽格式 ---
def pivot_results_to_wide_format(df_long: pd.DataFrame) -> pd.DataFrame:
    """
    将包含 '交易日' 和 '股票代码' 列的长格式 DataFrame
    转换为宽格式：每行一个交易日，列为 交易日, 股票代码1, 股票代码2, ...
    """
    if df_long.empty:
        # Return an empty DataFrame with the expected structure
        return pd.DataFrame(columns=['交易日'])

    # 1. 重置索引，为每个交易日内的股票创建序号
    df_reset = df_long.reset_index(drop=True)
    df_reset['row_id'] = df_reset.index

    # 2. 为每个交易日内的股票分配序号 (股票1, 股票2, ...)
    df_reset['stock_order'] = df_reset.groupby('交易日')['row_id'].rank(method='first').astype(int)
    
    # 3. 创建新的列名 '股票代码X'
    df_reset['stock_col'] = '股票代码' + df_reset['stock_order'].astype(str)

    # 4. 透视 (Pivot) 数据
    df_wide = df_reset.pivot(index='交易日', columns='stock_col', values='股票代码')
    
    # 5. 重置索引，使 '交易日' 成为普通列
    df_wide.reset_index(inplace=True)
    
    # 6. 确保列的顺序：'交易日' 在前，然后是 '股票代码1', '股票代码2', ...
    # 获取所有 '股票代码X' 列并排序
    stock_cols = [col for col in df_wide.columns if col.startswith('股票代码')]
    stock_cols.sort(key=lambda x: int(x.replace('股票代码', ''))) # 按数字排序
    # 重新排列列
    new_column_order = ['交易日'] + stock_cols
    df_wide = df_wide[new_column_order]

    # 7. 可选：清理列名 (如果 pivot 后有 '股票代码' 作为名称)
    # df_wide.columns.name = None # 如果需要，可以取消注释

    # print(f"数据已从长格式 ({len(df_long)} 行) 转换为宽格式 ({len(df_wide)} 行)。")
    return df_wide

# --- 辅助函数：设置特定工作表的列宽 ---
def set_column_widths(file_path: str, sheet_name: str, width_dict: dict):
    """
    设置 Excel 文件中指定工作表的列宽。
    :param file_path: Excel 文件路径
    :param sheet_name: 工作表名称
    :param width_dict: 列标识符到宽度的字典，例如 {'A': 8.38, 'B': 10.25}
    """
    try:
        # 加载工作簿
        wb = load_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            print(f"  警告：工作表 '{sheet_name}' 不存在于文件 '{file_path}' 中，无法设置列宽。")
            wb.close()
            return

        ws = wb[sheet_name]
        
        # 设置列宽
        for col_letter, width in width_dict.items():
            ws.column_dimensions[col_letter].width = width
            # print(f"  已设置工作表 '{sheet_name}' 列 '{col_letter}' 宽度为 {width}。")
        
        # 保存文件
        wb.save(file_path)
        wb.close()
        # print(f"  已更新文件 '{file_path}' 中工作表 '{sheet_name}' 的列宽。")
    except Exception as e:
        print(f"  设置工作表 '{sheet_name}' 列宽时出错: {e}")

# --- 主流程 ---
def main():
    """主函数"""
    print("--- 开始执行采样脚本 ---")
    
    # 1. 读取回测配置文件
    bt_config = get_sampling_config('sampling_config.txt')
    start_date = bt_config.get('start_date')
    end_date = bt_config.get('end_date')
    perform_filter = bt_config.get('filter', False)
    backtesting_mode = bt_config.get('BackTestingMode', 1) # 默认模式1
    
    if not start_date or not end_date:
        print("错误：配置文件 sampling_config.txt 中缺少 start_date 或 end_date。")
        return

    print(f"采样时间范围: {start_date} 至 {end_date}")
    print(f"是否执行二次筛选: {perform_filter}")
    print(f"回测模式: {backtesting_mode}")

    # 2. 建立交易日历 (复用 flu2.py 的函数)
    print("正在获取交易日历...")
    # 创建 ClickHouse 客户端 (复用 flu2.py 的函数)
    try:
        client = flu2.get_clickhouse_client()
    except Exception as e:
        print(f"无法连接到 ClickHouse 数据库: {e}")
        return
        
    # 调用 flu2.py 的 fetch_trading_calendar 函数
    try:
        trading_calendar = flu2.fetch_trading_calendar(client, start_date, end_date)
    except Exception as e:
        print(f"获取交易日历失败: {e}")
        return
        
    if not trading_calendar:
        print("错误：无法获取指定时间范围内的交易日历。")
        return
    print(f"获取到 {len(trading_calendar)} 个交易日。")

    # 3. 创建采样结果文件 sampling_result.xlsx (初始化)
    output_file = 'sampling_result.xlsx'
    print(f"初始化采样结果文件: {output_file}")
    # 初始化 Excel 文件和 sheet
    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # 创建空的 sheet，使用宽格式的列名初始化
            # 初始化时可以只写入列名
            pd.DataFrame(columns=['交易日']).to_excel(writer, sheet_name='samplingresults', index=False)
            pd.DataFrame(columns=['交易日']).to_excel(writer, sheet_name='filteredresults', index=False)
            # 创建 calendar sheet
            calendar_df = pd.DataFrame({'交易日': trading_calendar})
            calendar_df.to_excel(writer, sheet_name='calendar', index=False, startrow=0)
            # 写入交易日总数到 B1 (第1行, 第2列)
            worksheet = writer.sheets['calendar']
            worksheet.cell(row=1, column=2, value=len(trading_calendar)) # B1 is row=1, col=2
        print(f"采样结果文件 {output_file} 初始化完成。")
    except Exception as e:
        print(f"初始化采样结果文件 {output_file} 时出错: {e}")
        return

    # 4. 迭代执行策略
    all_results_data = [] # 用于存储所有原始结果 (长格式: 交易日, 股票代码)
    strategy_script = 'flu2.py'
    config_file = 'config.txt'
    temp_result_file = 'result.xlsx'

    for i, trade_date in enumerate(trading_calendar):
        print(f"\n--- 迭代 {i+1}/{len(trading_calendar)}: 处理交易日 {trade_date} ---")
        
        # 4.1 & 4.2: 更新配置文件
        if not update_config_file(config_file, trade_date):
            print(f"  跳过交易日 {trade_date}，因配置文件更新失败。")
            # 即使跳过，也添加一个空记录
            empty_day_df = pd.DataFrame({'交易日': [trade_date], '股票代码': [None]})
            all_results_data.append(empty_day_df)
            continue

        # 4.3: 执行策略脚本
        if not run_strategy_script(strategy_script):
            print(f"  跳过交易日 {trade_date}，因策略脚本执行失败。")
            # 即使失败也尝试删除可能产生的残余文件
            if os.path.exists(temp_result_file):
                try:
                    os.remove(temp_result_file)
                    print(f"  已删除残余结果文件 {temp_result_file}。")
                except Exception as e:
                    print(f"  删除残余结果文件 {temp_result_file} 时出错: {e}")
            # 添加一个空记录
            empty_day_df = pd.DataFrame({'交易日': [trade_date], '股票代码': [None]})
            all_results_data.append(empty_day_df)
            continue

        # 读取并处理结果
        df_result = read_strategy_results(temp_result_file, 'results')
        if not df_result.empty:
            # 添加交易日列
            df_result.insert(0, '交易日', trade_date)
            all_results_data.append(df_result)
            print(f"  交易日 {trade_date}: 成功获取 {len(df_result)} 条选股结果。")
        else:
            # 即使没有选出股票，也记录日期（创建一个只包含日期的行）
            empty_day_df = pd.DataFrame({'交易日': [trade_date], '股票代码': [None]})
            all_results_data.append(empty_day_df)
            print(f"  交易日 {trade_date}: 策略未选出股票。")

        # 4.4: 删除临时结果文件
        if os.path.exists(temp_result_file):
            try:
                os.remove(temp_result_file)
                print(f"  已删除临时结果文件 {temp_result_file}。")
            except Exception as e:
                print(f"  删除临时结果文件 {temp_result_file} 时出错: {e}")
        else:
            print(f"  临时结果文件 {temp_result_file} 不存在。")

    # 5. 将所有迭代结果转换格式并保存到 samplingresults sheet
    print("\n--- 处理并保存所有迭代结果到 samplingresults (宽格式) ---")
    if all_results_data:
        # 合并所有长格式数据
        df_all_results_long = pd.concat(all_results_data, ignore_index=True)
        
        # 转换为宽格式
        df_all_results_wide = pivot_results_to_wide_format(df_all_results_long)

        # 以追加模式打开文件并替换 sheet
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df_all_results_wide.to_excel(writer, sheet_name='samplingresults', index=False)
            print("所有迭代结果已按宽格式保存到 'samplingresults' sheet。")
            
            # 6. 执行二次筛选并将结果保存
            print("\n--- 处理筛选结果 ---")
            if perform_filter:
                df_filtered_results_long = filter_results(df_all_results_long, trading_calendar)
                # print("二次筛选完成，正在转换格式...")
            else:
                df_filtered_results_long = df_all_results_long.copy()
                print("未启用筛选，直接复制所有结果并转换格式。")
            
            # 将筛选后的长格式数据也转换为宽格式
            df_filtered_results_wide = pivot_results_to_wide_format(df_filtered_results_long)
            
            # 保存筛选后的宽格式数据
            with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                 df_filtered_results_wide.to_excel(writer, sheet_name='filteredresults', index=False)
            # print("筛选后的结果已按宽格式保存到 'filteredresults' sheet。")

            # --- 新增：设置列宽 ---
            # print("\n--- 设置 Excel 列宽 ---")
            # 定义列宽设置
            # 为 samplingresults 和 filteredresults 设置相同的列宽
            sheets_to_adjust = ['samplingresults', 'filteredresults']
            # 生成列宽字典: A: 8.38, B-U: 10.25
            # openpyxl 使用字母标识列
            column_widths = {'A': 8.38}
            # 添加 B 到 U 列的宽度 (B=2, U=21)
            for i in range(2, 22): # range(2, 22) 生成 2 到 21
                col_letter = chr(ord('A') + i - 1) # 2->B, 3->C, ..., 21->U
                column_widths[col_letter] = 10.25
            
            # 对每个需要调整的工作表应用列宽
            for sheet in sheets_to_adjust:
                 set_column_widths(output_file, sheet, column_widths)
            # print("Excel 列宽设置完成。")
            # --- 结束：设置列宽 ---

        except Exception as e:
            print(f"保存最终结果到 {output_file} 时出错: {e}")
            import traceback
            traceback.print_exc()
    else:
        print("未获取到任何选股结果。")

    print("\n--- 采样脚本执行完成 ---")

if __name__ == '__main__':
    main()
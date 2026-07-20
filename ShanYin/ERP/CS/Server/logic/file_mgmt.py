import os
import io
import pandas as pd
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Protection
from openpyxl.formatting.rule import FormulaRule
import traceback

from models import ChannelCustomer, Supplier, Point, SKU, ExternalPartner, BankAccount
from logic.constants import PointType, SupplierCategory, SKUType, ExternalPartnerType, AccountOwnerType, BankInfoKey
from logic.master.schemas import (
    CustomerSchema, PointSchema, SupplierSchema, SKUSchema, PartnerSchema, DeleteMasterDataSchema,
)
from logic.finance.schemas import CreateBankAccountSchema, UpdateBankAccountSchema
from logic.master.actions import (
    create_customer_action, update_customers_action, delete_customers_action,
    create_supplier_action, update_suppliers_action, delete_suppliers_action,
    create_partner_action, update_partners_action, delete_partners_action,
    create_point_action, update_points_action, delete_points_action,
    create_sku_action, update_skus_action, delete_skus_action,
)
from logic.finance.actions import create_bank_account_action, update_bank_accounts_action

def apply_protection_and_formatting(ws, df, op_col_letter, max_rows=10000):
    ws.protection.sheet = True # 开启工作表保护，默认全锁定
    ws.protection.formatColumns = False # 允许用户自由调整列宽
    
    # 将默认列宽拉宽一点，以防字被盖住
    for col_idx in range(1, len(df.columns) + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        col_name = df.columns[col_idx-1]
        
        if col_name in ["[系统ID]", "[操作指令]"]:
            ws.column_dimensions[col_letter].width = 10
        else:
            ws.column_dimensions[col_letter].width = 22
        
    # 性能优化：动态计算解锁范围
    # 单元格解锁( cell.protection) 是大批量操作时的性能瓶颈。
    # 我们只解锁“已有数据行 + 500行预留缓冲区”，确保日常操作秒开。
    # 下拉列表和条件格式依然覆盖 MAX_ROWS (10000)，不受此影响。
    unlock_max = min(max_rows, len(df) + 500)
    unlock_max = max(unlock_max, 1000) # 哪怕是空表，也至少解锁 1000 行以便录入
    
    max_col = len(df.columns)
    for row in ws.iter_rows(min_row=2, max_row=unlock_max, min_col=2, max_col=max_col):
        for cell in row:
            cell.protection = Protection(locked=False)
            
            # 将"银行账号"列格式显式设置为文本，防止导入/编辑时科学计数法截断
            if df.columns[cell.column - 1] == "银行账号":
                cell.number_format = '@'
    
    fmt_range = f"A2:{op_col_letter}{max_rows}"
    
    # 条件格式：选"删除"时整行变红
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    rule_delete = FormulaRule(formula=[f'${op_col_letter}2="删除"'], stopIfTrue=True, fill=red_fill)
    ws.conditional_formatting.add(fmt_range, rule_delete)
    
    # 条件格式：选"更新"时整行变蓝
    blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    rule_update = FormulaRule(formula=[f'${op_col_letter}2="更新"'], stopIfTrue=True, fill=blue_fill)
    ws.conditional_formatting.add(fmt_range, rule_update)

def generate_master_data_excel(session) -> bytes:
    """导出最新的数据库中所有基础资料为带验证规则的 Excel Bytes"""
    MAX_ROWS = 10000
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # 预先创建隐藏配置页，供后续级联下拉引用
        ws_settings = writer.book.create_sheet('Settings')
        ws_settings.sheet_state = 'hidden'
        
        # 预加载所有实体名称
        all_customers = session.query(ChannelCustomer).all()
        all_suppliers = session.query(Supplier).all()
        all_partners = session.query(ExternalPartner).all()
        
        # A列：公司自身
        ws_settings['A1'] = "闪饮自身"
        # B列：客户名称列表
        for i, c in enumerate(all_customers):
            ws_settings.cell(row=i+1, column=2, value=c.name)
        # C列：供应商名称列表
        for i, s in enumerate(all_suppliers):
            ws_settings.cell(row=i+1, column=3, value=s.name)
        # D列：合作方名称列表
        for i, p in enumerate(all_partners):
            ws_settings.cell(row=i+1, column=4, value=p.name)

        # ==========================================
        # 1. 渠道客户 
        # ==========================================
        customers = session.query(ChannelCustomer).all()
        cust_data = []
        for c in customers:
            cust_data.append({
                "[系统ID]": c.id,
                "客户名称": c.name,
                "整体信息描述": c.info,
                "[操作指令]": ""
            })
        df_cust = pd.DataFrame(cust_data) if cust_data else pd.DataFrame(columns=["[系统ID]", "客户名称", "整体信息描述", "[操作指令]"])
        df_cust.to_excel(writer, sheet_name="渠道客户", index=False)
        
        ws_cust = writer.sheets["渠道客户"]
        dv_op_cust = DataValidation(type="list", formula1='"新增,更新,删除"', allow_blank=True, showErrorMessage=True)
        ws_cust.add_data_validation(dv_op_cust)
        dv_op_cust.add(f"D2:D{MAX_ROWS}")
        apply_protection_and_formatting(ws_cust, df_cust, "D")

        # ==========================================
        # 2. 供应商 
        # ==========================================
        suppliers = session.query(Supplier).all()
        supp_data = []
        for s in suppliers:
            supp_data.append({
                "[系统ID]": s.id,
                "供应商名称": s.name,
                "供应类别": s.category,
                "地址信息": s.address,
                "[操作指令]": ""
            })
        df_supp = pd.DataFrame(supp_data) if supp_data else pd.DataFrame(columns=["[系统ID]", "供应商名称", "供应类别", "地址信息", "[操作指令]"])
        df_supp.to_excel(writer, sheet_name="供应商", index=False)
        
        ws_supp = writer.sheets["供应商"]
        dv_supp_cat = DataValidation(type="list", formula1=f'"{",".join(SupplierCategory.ALL_TYPES)}"', allow_blank=True, showErrorMessage=True)
        ws_supp.add_data_validation(dv_supp_cat)
        dv_supp_cat.add(f"C2:C{MAX_ROWS}")
        dv_op_supp = DataValidation(type="list", formula1='"新增,更新,删除"', allow_blank=True, showErrorMessage=True)
        ws_supp.add_data_validation(dv_op_supp)
        dv_op_supp.add(f"E2:E{MAX_ROWS}")
        apply_protection_and_formatting(ws_supp, df_supp, "E")
        
        # ==========================================
        # 3. 点位
        # ==========================================
        points = session.query(Point).all()
        point_data = []
        for p in points:
            owner_type = "[自身] 公司"
            owner_name = "闪饮自身"
            if p.customer:
                owner_type = "[客户]"
                owner_name = p.customer.name
            elif p.supplier:
                owner_type = "[供应商]"
                owner_name = p.supplier.name
                
            point_data.append({
                "[系统ID]": p.id,
                "归属主体类别": owner_type,
                "归属主体名称": owner_name,
                "点位名称": p.name,
                "点位类型": p.type,
                "详细地址": p.address,
                "收货地址": p.receiving_address,
                "[操作指令]": ""
            })
        df_point = pd.DataFrame(point_data) if point_data else pd.DataFrame(columns=["[系统ID]", "归属主体类别", "归属主体名称", "点位名称", "点位类型", "详细地址", "收货地址", "[操作指令]"])
        df_point.to_excel(writer, sheet_name="点位", index=False)
        
        ws_point = writer.sheets["点位"]
        dv_owner_type = DataValidation(type="list", formula1='"[自身] 公司,[客户],[供应商]"', allow_blank=True, showErrorMessage=True)
        ws_point.add_data_validation(dv_owner_type)
        dv_owner_type.add(f"B2:B{MAX_ROWS}")
        # 归属主体名称级联下拉 - 使用 INDIRECT 避免行偏移 bug
        # 点位只有客户和供应商，没有合作方
        dv_owner_name = DataValidation(type="list", 
                                      formula1=f'=INDIRECT(IF($B2="[客户]","Settings!$B$1:$B${MAX_ROWS}",IF($B2="[供应商]","Settings!$C$1:$C${MAX_ROWS}","Settings!$A$1:$A$1")))', 
                                      allow_blank=True, showErrorMessage=True)
        ws_point.add_data_validation(dv_owner_name)
        dv_owner_name.add(f"C2:C{MAX_ROWS}")
        dv_point_type = DataValidation(type="list", formula1=f'"{",".join(PointType.ALL_TYPES)}"', allow_blank=True, showErrorMessage=True)
        ws_point.add_data_validation(dv_point_type)
        dv_point_type.add(f"E2:E{MAX_ROWS}")
        dv_op_point = DataValidation(type="list", formula1='"新增,更新,删除"', allow_blank=True, showErrorMessage=True)
        ws_point.add_data_validation(dv_op_point)
        dv_op_point.add(f"H2:H{MAX_ROWS}")
        apply_protection_and_formatting(ws_point, df_point, "H")
        
        # ==========================================
        # 4. SKU
        # ==========================================
        skus = session.query(SKU).all()
        sku_data = []
        for k in skus:
            sku_data.append({
                "[系统ID]": k.id,
                "所属供应商名称": k.supplier.name if k.supplier else "",
                "SKU名称": k.name,
                "一级分类": k.type_level1,
                "型号": k.model,
                "[操作指令]": ""
            })
        df_sku = pd.DataFrame(sku_data) if sku_data else pd.DataFrame(columns=["[系统ID]", "所属供应商名称", "SKU名称", "一级分类", "型号", "[操作指令]"])
        df_sku.to_excel(writer, sheet_name="SKU", index=False)
        
        ws_sku = writer.sheets["SKU"]
        # SKU的供应商下拉也改用 INDIRECT 从 Settings 页取值
        dv_sku_supplier = DataValidation(type="list", formula1=f'=INDIRECT("Settings!$C$1:$C${MAX_ROWS}")', allow_blank=True, showErrorMessage=True)
        ws_sku.add_data_validation(dv_sku_supplier)
        dv_sku_supplier.add(f"B2:B{MAX_ROWS}")
        dv_sku_type = DataValidation(type="list", formula1=f'"{",".join(SKUType.ALL_TYPES)}"', allow_blank=True, showErrorMessage=True)
        ws_sku.add_data_validation(dv_sku_type)
        dv_sku_type.add(f"D2:D{MAX_ROWS}")
        dv_op_sku = DataValidation(type="list", formula1='"新增,更新,删除"', allow_blank=True, showErrorMessage=True)
        ws_sku.add_data_validation(dv_op_sku)
        dv_op_sku.add(f"F2:F{MAX_ROWS}")
        apply_protection_and_formatting(ws_sku, df_sku, "F")
        
        # ==========================================
        # 5. 外部合作方
        # ==========================================
        partners = session.query(ExternalPartner).all()
        partner_data = []
        for p in partners:
            partner_data.append({
                "[系统ID]": p.id,
                "机构名称": p.name,
                "机构类别": p.type,
                "[操作指令]": ""
            })
        df_partner = pd.DataFrame(partner_data) if partner_data else pd.DataFrame(columns=["[系统ID]", "机构名称", "机构类别", "[操作指令]"])
        df_partner.to_excel(writer, sheet_name="外部合作方", index=False)
        
        ws_partner = writer.sheets["外部合作方"]
        dv_partner_type = DataValidation(type="list", formula1=f'"{",".join(ExternalPartnerType.ALL_TYPES)}"', allow_blank=True, showErrorMessage=True)
        ws_partner.add_data_validation(dv_partner_type)
        dv_partner_type.add(f"C2:C{MAX_ROWS}")
        dv_op_partner = DataValidation(type="list", formula1='"新增,更新,删除"', allow_blank=True, showErrorMessage=True)
        ws_partner.add_data_validation(dv_op_partner)
        dv_op_partner.add(f"D2:D{MAX_ROWS}")
        apply_protection_and_formatting(ws_partner, df_partner, "D")
        
        # ==========================================
        # 6. 银行账户
        # ==========================================
        bank_accs = session.query(BankAccount).all()
        bank_data = []
        for ba in bank_accs:
            # Resolve nature and name
            nature = "[自身] 公司"
            name_val = "闪饮自身"
            if ba.owner_type == AccountOwnerType.CUSTOMER and ba.owner_id:
                c = session.query(ChannelCustomer).get(ba.owner_id)
                nature = "[客户]"
                name_val = c.name if c else "未知客户"
            elif ba.owner_type == AccountOwnerType.SUPPLIER and ba.owner_id:
                s = session.query(Supplier).get(ba.owner_id)
                nature = "[供应商]"
                name_val = s.name if s else "未知供应商"
            elif ba.owner_type == AccountOwnerType.PARTNER and ba.owner_id:
                p = session.query(ExternalPartner).get(ba.owner_id)
                nature = "[合作方]"
                name_val = p.name if p else "未知合作方"
            
            info = ba.account_info or {}
            bank_data.append({
                "[系统ID]": ba.id,
                "归属方性质": nature,
                "归属方名称": name_val,
                "开户名称": info.get(BankInfoKey.HOLDER_NAME, ""),
                "银行名称": info.get(BankInfoKey.BANK_NAME, ""),
                "银行账号": info.get(BankInfoKey.ACCOUNT_NO, ""),
                "默认账户": "是" if ba.is_default else "否",
                "[操作指令]": ""
            })
        df_bank = pd.DataFrame(bank_data) if bank_data else pd.DataFrame(columns=["[系统ID]", "归属方性质", "归属方名称", "开户名称", "银行名称", "银行账号", "默认账户", "[操作指令]"])
        df_bank.to_excel(writer, sheet_name="银行账户", index=False)
        
        ws_bank = writer.sheets["银行账户"]
        # 归属方性质下拉
        dv_owner_nature = DataValidation(type="list", formula1='"[自身] 公司,[客户],[供应商],[合作方]"', allow_blank=True, showErrorMessage=True)
        ws_bank.add_data_validation(dv_owner_nature)
        dv_owner_nature.add(f"B2:B{MAX_ROWS}")
        
        # 归属方名称级联下拉 - 使用 INDIRECT 避免行偏移 bug
        # 客户→Settings B列，供应商→Settings C列，合作方→Settings D列，自身→Settings A1
        dv_owner_name = DataValidation(type="list", 
                                      formula1=f'=INDIRECT(IF($B2="[客户]","Settings!$B$1:$B${MAX_ROWS}",IF($B2="[供应商]","Settings!$C$1:$C${MAX_ROWS}",IF($B2="[合作方]","Settings!$D$1:$D${MAX_ROWS}","Settings!$A$1:$A$1"))))', 
                                      allow_blank=True, showErrorMessage=True)
        ws_bank.add_data_validation(dv_owner_name)
        dv_owner_name.add(f"C2:C{MAX_ROWS}")

        # 默认账户和操作指令 (列后移了：默认在G，操作指令在H)
        dv_bank_default = DataValidation(type="list", formula1='"是,否"', allow_blank=True, showErrorMessage=True)
        ws_bank.add_data_validation(dv_bank_default)
        dv_bank_default.add(f"G2:G{MAX_ROWS}")
        dv_op_bank = DataValidation(type="list", formula1='"新增,更新,删除"', allow_blank=True, showErrorMessage=True)
        ws_bank.add_data_validation(dv_op_bank)
        dv_op_bank.add(f"H2:H{MAX_ROWS}")
        apply_protection_and_formatting(ws_bank, df_bank, "H")
        
    output.seek(0)
    return output.getvalue()

def process_master_data_excel(session, file_bytes: bytes) -> dict:
    """处理上传的Excel记录，同步到数据库"""
    report = {
        "success": True,
        "logs": [], # 记录具体的提示和报错字符串
        "stats": {"新增": 0, "更新": 0, "删除": 0}
    }
    
    try:
        xls = pd.ExcelFile(io.BytesIO(file_bytes), engine='openpyxl')
        expected_sheets = ["渠道客户", "供应商", "点位", "SKU", "外部合作方", "银行账户"]
        missing_sheets = [s for s in expected_sheets if s not in xls.sheet_names]
        if missing_sheets:
            report["success"] = False
            report["logs"].append(f"❌ 模板校验失败: 缺少页签 {', '.join(missing_sheets)}")
            return report
            
        # ==========================================
        # Phase 1: 无外键依赖的独立实体 (Customer, Supplier, Partner)
        # ==========================================
        
        # 1. 渠道客户 
        df_cust = pd.read_excel(xls, sheet_name="渠道客户")
        df_cust.fillna("", inplace=True)
        cust_creates, cust_updates, cust_deletes = [], [], []
        
        for idx, row in df_cust.iterrows():
            cid = row.get("[系统ID]")
            op = row.get("[操作指令]")
            name = str(row.get("客户名称", "")).strip()
            if not name and op != "删除": continue
            
            if op == "删除" and cid:
                cust_deletes.append(DeleteMasterDataSchema(id=int(cid)))
            elif op == "更新" and cid:
                cust_updates.append(CustomerSchema(id=int(cid), name=name, info=str(row.get("整体信息描述", ""))))
            elif (not cid or op == "新增") and name:
                res = create_customer_action(session, CustomerSchema(name=name, info=str(row.get("整体信息描述", ""))))
                if res.success: report["stats"]["新增"] += 1
                else: report["logs"].append(f"❌ 客户新增失败 [{name}]: {res.error}")

        if cust_updates:
            res = update_customers_action(session, cust_updates)
            if res.success: report["stats"]["更新"] += len(cust_updates)
            else: report["logs"].append(f"❌ 客户批量更新失败: {res.error}")
        if cust_deletes:
            res = delete_customers_action(session, cust_deletes)
            if res.success: report["stats"]["删除"] += len(cust_deletes)
            else: report["logs"].append(f"❌ 客户删除失败: {res.error}")

        # 2. 供应商
        df_supp = pd.read_excel(xls, sheet_name="供应商")
        df_supp.fillna("", inplace=True)
        supp_creates, supp_updates, supp_deletes = [], [], []

        for idx, row in df_supp.iterrows():
            sid = row.get("[系统ID]")
            op = row.get("[操作指令]")
            name = str(row.get("供应商名称", "")).strip()
            if not name and op != "删除": continue
            
            if op == "删除" and sid:
                supp_deletes.append(DeleteMasterDataSchema(id=int(sid)))
            elif op == "更新" and sid:
                supp_updates.append(SupplierSchema(id=int(sid), name=name, category=str(row.get("供应类别", "")), address=str(row.get("地址信息", ""))))
            elif (not sid or op == "新增") and name:
                res = create_supplier_action(session, SupplierSchema(name=name, category=str(row.get("供应类别", "")), address=str(row.get("地址信息", ""))))
                if res.success: report["stats"]["新增"] += 1
                else: report["logs"].append(f"❌ 供应商新增失败 [{name}]: {res.error}")
                
        if supp_updates:
            res = update_suppliers_action(session, supp_updates)
            if res.success: report["stats"]["更新"] += len(supp_updates)
            else: report["logs"].append(f"❌ 供应商批量更新失败: {res.error}")
        if supp_deletes:
            res = delete_suppliers_action(session, supp_deletes)
            if res.success: report["stats"]["删除"] += len(supp_deletes)
            else: report["logs"].append(f"❌ 供应商删除失败: {res.error}")

        # 3. 外部合作方
        df_partner = pd.read_excel(xls, sheet_name="外部合作方")
        df_partner.fillna("", inplace=True)
        part_creates, part_updates, part_deletes = [], [], []

        for idx, row in df_partner.iterrows():
            pid = row.get("[系统ID]")
            op = row.get("[操作指令]")
            name = str(row.get("机构名称", "")).strip()
            if not name and op != "删除": continue
            
            if op == "删除" and pid:
                part_deletes.append(DeleteMasterDataSchema(id=int(pid)))
            elif op == "更新" and pid:
                part_updates.append(PartnerSchema(id=int(pid), name=name, type=str(row.get("机构类别", ""))))
            elif (not pid or op == "新增") and name:
                res = create_partner_action(session, PartnerSchema(name=name, type=str(row.get("机构类别", ""))))
                if res.success: report["stats"]["新增"] += 1
                else: report["logs"].append(f"❌ 合作方新增失败 [{name}]: {res.error}")

        if part_updates:
            res = update_partners_action(session, part_updates)
            if res.success: report["stats"]["更新"] += len(part_updates)
            else: report["logs"].append(f"❌ 合作方批量更新失败: {res.error}")
        if part_deletes:
            res = delete_partners_action(session, part_deletes)
            if res.success: report["stats"]["删除"] += len(part_deletes)
            else: report["logs"].append(f"❌ 合作方删除失败: {res.error}")

        # 刷新数据库状态，以便构建字典用于下一级的FK翻译
        session.flush()

        # 构建最新字典映射
        customer_map = {c.name.strip(): c.id for c in session.query(ChannelCustomer).all()}
        supplier_map = {s.name.strip(): s.id for s in session.query(Supplier).all()}

        # ==========================================
        # Phase 2: 有外键依赖的实体 (Point, SKU)
        # ==========================================

        # 4. SKU
        df_sku = pd.read_excel(xls, sheet_name="SKU")
        df_sku.fillna("", inplace=True)
        sku_creates, sku_updates, sku_deletes = [], [], []

        for idx, row in df_sku.iterrows():
            kid = row.get("[系统ID]")
            op = row.get("[操作指令]")
            name = str(row.get("SKU名称", "")).strip()
            supp_name = str(row.get("所属供应商名称", "")).strip()
            
            if op == "删除" and kid:
                sku_deletes.append(DeleteMasterDataSchema(id=int(kid)))
                continue
                
            if not name or not supp_name: continue
            
            mapped_supp_id = supplier_map.get(supp_name)
            if not mapped_supp_id:
                report["logs"].append(f"⚠️ 第 {idx+2} 行 SKU [{name}] 处理跳过：找不到指定的供应商名称 '{supp_name}'")
                continue
            
            payload = SKUSchema(
                id=int(kid) if kid else None,
                supplier_id=mapped_supp_id,
                name=name,
                type_level1=str(row.get("一级分类", "")),
                model=str(row.get("型号", ""))
            )
            
            if op == "更新" and kid:
                sku_updates.append(payload)
            elif not kid or op == "新增":
                res = create_sku_action(session, payload)
                if res.success: report["stats"]["新增"] += 1
                else: report["logs"].append(f"❌ SKU新增失败 [{name}]: {res.error}")

        if sku_updates:
            res = update_skus_action(session, sku_updates)
            if res.success: report["stats"]["更新"] += len(sku_updates)
            else: report["logs"].append(f"❌ SKU批量更新失败: {res.error}")
        if sku_deletes:
            res = delete_skus_action(session, sku_deletes)
            if res.success: report["stats"]["删除"] += len(sku_deletes)
            else: report["logs"].append(f"❌ SKU删除失败: {res.error}")
            
        # 5. 点位
        df_point = pd.read_excel(xls, sheet_name="点位")
        df_point.fillna("", inplace=True)
        pt_creates, pt_updates, pt_deletes = [], [], []

        for idx, row in df_point.iterrows():
            ptid = row.get("[系统ID]")
            op = row.get("[操作指令]")
            name = str(row.get("点位名称", "")).strip()
            owner_base = str(row.get("归属主体类别", "")).strip()
            owner_name = str(row.get("归属主体名称", "")).strip()
            
            if op == "删除" and ptid:
                pt_deletes.append(DeleteMasterDataSchema(id=int(ptid)))
                continue
                
            if not name or not owner_base: continue
            
            mapped_c_id, mapped_s_id = None, None
            if owner_base == "[客户]":
                mapped_c_id = customer_map.get(owner_name)
                if not mapped_c_id:
                    report["logs"].append(f"⚠️ 第 {idx+2} 行点位 [{name}] 处理跳过：找不到客户 '{owner_name}'")
                    continue
            elif owner_base == "[供应商]":
                mapped_s_id = supplier_map.get(owner_name)
                if not mapped_s_id:
                    report["logs"].append(f"⚠️ 第 {idx+2} 行点位 [{name}] 处理跳过：找不到供应商 '{owner_name}'")
                    continue
            
            payload = PointSchema(
                id=int(ptid) if ptid else None,
                name=name,
                customer_id=mapped_c_id,
                supplier_id=mapped_s_id,
                type=str(row.get("点位类型", "")),
                address=str(row.get("详细地址", "")),
                receiving_address=str(row.get("收货地址", ""))
            )
            
            if op == "更新" and ptid:
                pt_updates.append(payload)
            elif not ptid or op == "新增":
                res = create_point_action(session, payload)
                if res.success: report["stats"]["新增"] += 1
                else: report["logs"].append(f"❌ 点位新增失败 [{name}]: {res.error}")

        if pt_updates:
            res = update_points_action(session, pt_updates)
            if res.success: report["stats"]["更新"] += len(pt_updates)
            else: report["logs"].append(f"❌ 点位批量更新失败: {res.error}")
        if pt_deletes:
            res = delete_points_action(session, pt_deletes)
            if res.success: report["stats"]["删除"] += len(pt_deletes)
            else: report["logs"].append(f"❌ 点位删除失败: {res.error}")

        # ==========================================
        # Phase 3: 银行账户 (依赖客户/供应商/合作方的名字映射)
        # 规则：每个主体必须有且仅有一个默认账户，否则该主体所有账户行全部拒绝处理
        # ==========================================
        from collections import defaultdict
        partner_map = {p.name.strip(): p.id for p in session.query(ExternalPartner).all()}
        
        df_bank = pd.read_excel(xls, sheet_name="银行账户")
        df_bank.fillna("", inplace=True)
        
        # --- 第一遍：解析所有行，按主体分组 ---
        parsed_rows = []  # list of dicts with resolved info
        owner_groups = defaultdict(list)  # owner_key -> [parsed_row_indices]
        
        for idx, row in df_bank.iterrows():
            bid = row.get("[系统ID]")
            op = str(row.get("[操作指令]", "")).strip()
            owner_nature = str(row.get("归属方性质", "")).strip()
            owner_name = str(row.get("归属方名称", "")).strip()
            acc_name = str(row.get("开户名称", "")).strip()
            bank_name_val = str(row.get("银行名称", "")).strip()
            acc_num = str(row.get("银行账号", "")).strip()
            is_def = str(row.get("默认账户", "否")).strip() == "是"
            
            if not owner_nature and op != "删除": continue
            
            # Resolve owner_type and owner_id from label
            resolved_type, resolved_id = None, None
            resolve_ok = True
            if "公司" in owner_nature or "自身" in owner_nature:
                resolved_type = AccountOwnerType.OURSELVES
                resolved_id = None
            elif owner_nature == "[客户]":
                resolved_id = customer_map.get(owner_name)
                if not resolved_id:
                    report["logs"].append(f"⚠️ 第 {idx+2} 行银行账户处理跳过：找不到客户 '{owner_name}'")
                    resolve_ok = False
                resolved_type = AccountOwnerType.CUSTOMER
            elif owner_nature == "[供应商]":
                resolved_id = supplier_map.get(owner_name)
                if not resolved_id:
                    report["logs"].append(f"⚠️ 第 {idx+2} 行银行账户处理跳过：找不到供应商 '{owner_name}'")
                    resolve_ok = False
                resolved_type = AccountOwnerType.SUPPLIER
            elif owner_nature == "[合作方]":
                resolved_id = partner_map.get(owner_name)
                if not resolved_id:
                    report["logs"].append(f"⚠️ 第 {idx+2} 行银行账户处理跳过：找不到合作方 '{owner_name}'")
                    resolve_ok = False
                resolved_type = AccountOwnerType.PARTNER
            else:
                report["logs"].append(f"⚠️ 第 {idx+2} 行银行账户处理跳过：无法识别归属方性质 '{owner_nature}'")
                resolve_ok = False
            
            if not resolve_ok: continue
                
            parsed = {
                "excel_row": idx + 2,
                "bid": bid, "op": op,
                "owner_label": f"{owner_nature}{owner_name}",
                "resolved_type": resolved_type, "resolved_id": resolved_id,
                "account_info": {BankInfoKey.HOLDER_NAME: acc_name, BankInfoKey.BANK_NAME: bank_name_val, BankInfoKey.ACCOUNT_NO: acc_num},
                "is_default": is_def
            }
            row_index = len(parsed_rows)
            parsed_rows.append(parsed)
            
            # 按主体分组 (owner_type, owner_id) 作为 key
            owner_key = (resolved_type, resolved_id)
            owner_groups[owner_key].append(row_index)
        
        # --- 第二遍：校验每个主体的默认账户数量 ---
        blocked_owners = set()
        for owner_key, row_indices in owner_groups.items():
            default_count = sum(1 for ri in row_indices if parsed_rows[ri]["is_default"])
            total_count = len(row_indices)
            
            if default_count == 0:
                owner_label = parsed_rows[row_indices[0]]["owner_label"]
                report["logs"].append(f"❌ 归属方 [{owner_label}] 的 {total_count} 个银行账户中没有设置默认账户，全部跳过。请确保恰好有一个标记为\"是\"。")
                blocked_owners.add(owner_key)
            elif default_count > 1:
                owner_label = parsed_rows[row_indices[0]]["owner_label"]
                report["logs"].append(f"❌ 归属方 [{owner_label}] 的 {total_count} 个银行账户中有 {default_count} 个被标记为默认，全部跳过。请确保恰好有一个标记为\"是\"。")
                blocked_owners.add(owner_key)
        
        # --- 第三遍：只对通过校验的主体执行入库 ---
        for p in parsed_rows:
            owner_key = (p["resolved_type"], p["resolved_id"])
            if owner_key in blocked_owners:
                continue
            
            bid = p["bid"]
            op = p["op"]

            if op == "删除" and bid:
                obj = session.query(BankAccount).get(int(bid))
                if obj:
                    session.delete(obj)
                    session.commit()
                    report["stats"]["删除"] += 1
                continue
                
            if op == "更新" and bid:
                res = update_bank_accounts_action(session, [UpdateBankAccountSchema(
                    id=int(bid), owner_type=p["resolved_type"], owner_id=p["resolved_id"],
                    account_info=p["account_info"], is_default=p["is_default"]
                )])
                if res.success: report["stats"]["更新"] += 1
                else: report["logs"].append(f"❌ 银行账户更新失败 (ID={bid}): {res.error}")
            elif not bid or op == "新增":
                res = create_bank_account_action(session, CreateBankAccountSchema(
                    owner_type=p["resolved_type"], owner_id=p["resolved_id"],
                    account_info=p["account_info"], is_default=p["is_default"]
                ))
                if res.success: report["stats"]["新增"] += 1
                else: report["logs"].append(f"❌ 银行账户新增失败: {res.error}")
            
    except Exception as e:
        report["success"] = False
        report["logs"].append(f"❌ 严重错误: {str(e)}")
        report["logs"].append(traceback.format_exc())
        
    return report

# --- 合同附件管理 ---
CONTRACT_DIR = "data/contracts"

def save_contract_files(contract_id, uploaded_files):
    """保存合同附件到本地目录"""
    target_dir = os.path.join(CONTRACT_DIR, str(contract_id))
    os.makedirs(target_dir, exist_ok=True)
    saved = []
    for f in uploaded_files:
        path = os.path.join(target_dir, f.name)
        with open(path, "wb") as out:
            out.write(f.getbuffer())
        saved.append(f.name)
    return saved

def get_contract_files(contract_id):
    """获取指定合同的附件列表"""
    target_dir = os.path.join(CONTRACT_DIR, str(contract_id))
    if not os.path.exists(target_dir):
        return []
    return [f for f in os.listdir(target_dir) if os.path.isfile(os.path.join(target_dir, f))]


# --- 批次质检报告管理 ---
BATCH_CERT_DIR = "data/batchcertificate"

def save_batch_certificate(batch_no: str, uploaded_file) -> str:
    """保存批次质检报告到本地目录，返回保存后的文件路径"""
    os.makedirs(BATCH_CERT_DIR, exist_ok=True)
    ext = os.path.splitext(uploaded_file.filename)[1] if uploaded_file.filename else ""
    filename = f"{batch_no}{ext}"
    path = os.path.join(BATCH_CERT_DIR, filename)
    with open(path, "wb") as out:
        out.write(uploaded_file.file.read())
    return path


def get_batch_certificate_path(batch_no: str) -> str | None:
    """根据批次号查找质检报告路径（支持多种扩展名）"""
    if not os.path.exists(BATCH_CERT_DIR):
        return None
    for f in os.listdir(BATCH_CERT_DIR):
        if f.startswith(batch_no):
            return os.path.join(BATCH_CERT_DIR, f)
    return None
